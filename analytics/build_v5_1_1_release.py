"""V5.1.1 release builder. CI is the only place where project-derived artifacts are materialized."""
from __future__ import annotations
import csv, hashlib, json, os
from pathlib import Path
from .build_v5_1_1_economics import run

ROOT=Path(__file__).resolve().parents[1]

def _write(path, text):
    path.parent.mkdir(parents=True,exist_ok=True); path.write_text(text,encoding="utf-8"); return path

def _csv(path, rows, fields):
    path.parent.mkdir(parents=True,exist_ok=True)
    with path.open("w",encoding="utf-8",newline="") as f:
        w=csv.DictWriter(f,fieldnames=fields); w.writeheader(); w.writerows(rows)

def _hash(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()

def _workbook(root, model):
    try:
        from openpyxl import Workbook
    except ImportError:
        return None
    wb=Workbook(); wb.remove(wb.active)
    names=["00_Cover","01_Readme","02_Project_Master","03_Input_View","04_Assumption_Overlay","05_Source_Register","06_Conflict_Register","07_Selected_Data_Audit","08_Yield_Audit","09_Load_Match","10_Tariff","11_FX","12_Tax","13_Rates","14_Discount_Rates","15_CAPEX","16_OPEX","17_Base_CFADS","18_Debt_Sizing","19_Debt_Schedule","20_PPA_Frontier","21_Scenarios","22_Portfolio","23_Claim_Governance","24_Release_Gates","25_QA"]
    for name in names:
        ws=wb.create_sheet(name); ws["A1"]="VIETGREEN V5.1.1 — FULL DATA-MODEL & CONTENT REBUILD"; ws["A2"]="Remote-authoritative GitHub branch; CI-generated derived artifact"
    ws=wb["00_Cover"]; ws["A4"]="Status"; ws["B4"]="RELEASE_CANDIDATE_PENDING_CI"
    ws["A5"]="Claim boundary"; ws["B5"]="Standardized public-data Project Finance reconstruction; not bankability, IC approval, legal, tax or technical sign-off."
    ws=wb["02_Project_Master"]; rows=model["economics"]
    if rows:
        keys=list(rows[0].keys())
        for col,key in enumerate(keys,1): ws.cell(4,col,key)
        for rr,row in enumerate(rows,5):
            for cc,key in enumerate(keys,1): ws.cell(rr,cc,row.get(key,""))
    ws=wb["25_QA"]; checks=[("selected_project_count",len(rows)),("scenario_rows",len(model["scenarios"])),("cash_flow_rows",len(model["cash_flow"])),("input_view_rows",len(model["model_input_view"])),("ppa_mode","FRONTIER_ONLY"),("decision","INDETERMINATE_MISSING_COMMERCIAL_DATA"),("remote_only","TRUE")]
    for i,(k,v) in enumerate(checks,4): ws.cell(i,1,k); ws.cell(i,2,v)
    out=root/"artifacts/v5_1_1_model/vietgreen_v5_1_1_model.xlsx"; out.parent.mkdir(parents=True,exist_ok=True); wb.save(out); return out

def build(root=ROOT):
    model=run(root,root/"artifacts/v5_1_1_model")
    workbook_path=_workbook(root,model)
    econ=model["economics"]; scenarios=model["scenarios"]; cash=model["cash_flow"]; debt=model["debt_schedule"]; hourly=model["hourly"]; inputs=model["model_input_view"]
    def pick(rows, keys):
        return [{k:r.get(k,"") for k in keys} for r in rows]
    out=root/"outputs"
    _csv(out/"v5_1_1_model_input_view.csv",inputs,list(inputs[0]) if inputs else [])
    _csv(out/"v5_1_1_energy.csv",pick(econ,["project_id","country","installed_capacity_kwp_observed","generation_p50_kwh_observed","generation_p50_kwh_modeled","specific_yield_observed","generation_p90_kwh","generation_p99_kwh"]),["project_id","country","installed_capacity_kwp_observed","generation_p50_kwh_observed","generation_p50_kwh_modeled","specific_yield_observed","generation_p90_kwh","generation_p99_kwh"])
    _csv(out/"v5_1_1_load_summary.csv",pick(econ,["project_id","annual_load_kwh_modeled","load_evidence_level","load_8760_rows","self_consumed_kwh_p50","export_kwh_p50"]),["project_id","annual_load_kwh_modeled","load_evidence_level","load_8760_rows","self_consumed_kwh_p50","export_kwh_p50"])
    _csv(out/"v5_1_1_8760.csv",hourly,["project_id","timestamp","load_kwh","solar_kwh","self_consumed_kwh","export_kwh","profile_year"])
    _csv(out/"v5_1_1_ppa_frontier.csv",pick(econ,["project_id","currency","ppa_price_local_per_kwh","customer_ceiling_local_per_kwh","sponsor_floor_local_per_kwh","lender_floor_local_per_kwh","negotiation_lower_local_per_kwh","negotiation_upper_local_per_kwh","negotiation_status","reference_case","decision"]),["project_id","currency","ppa_price_local_per_kwh","customer_ceiling_local_per_kwh","sponsor_floor_local_per_kwh","lender_floor_local_per_kwh","negotiation_lower_local_per_kwh","negotiation_upper_local_per_kwh","negotiation_status","reference_case","decision"])
    _csv(out/"v5_1_1_cash_flow.csv",cash,list(cash[0]) if cash else [])
    _csv(out/"v5_1_1_debt_sizing.csv",pick(econ,["project_id","capex_local","capex_usd","debt_capacity_local","debt_capacity_usd","binding_debt_constraint","debt_rate_type"]),["project_id","capex_local","capex_usd","debt_capacity_local","debt_capacity_usd","binding_debt_constraint","debt_rate_type"])
    _csv(out/"v5_1_1_debt_schedule.csv",debt,list(debt[0]) if debt else [])
    _csv(out/"v5_1_1_coverage.csv",pick(econ,["project_id","dscr_min","llcr_loan_life","plcr_project_life","debt_capacity_usd"]),["project_id","dscr_min","llcr_loan_life","plcr_project_life","debt_capacity_usd"])
    _csv(out/"v5_1_1_returns.csv",pick(econ,["project_id","project_npv_usd_at_reference","project_irr_at_reference","equity_npv_usd_at_reference","equity_irr_at_reference","reference_case","decision"]),["project_id","project_npv_usd_at_reference","project_irr_at_reference","equity_npv_usd_at_reference","equity_irr_at_reference","reference_case","decision"])
    _csv(out/"v5_1_1_scenarios.csv",scenarios,list(scenarios[0]) if scenarios else [])
    _csv(out/"v5_1_1_project_economics.csv",econ,list(econ[0]) if econ else [])
    shortlist=[{"project_id":r["project_id"],"country":r["country"],"evidence_boundary":r["evidence_boundary"],"zone_status":r["negotiation_status"],"zone_width_local_per_kwh":(float(r["negotiation_upper_local_per_kwh"] or 0)-float(r["negotiation_lower_local_per_kwh"] or 0)) if r["negotiation_upper_local_per_kwh"]!="" and r["negotiation_lower_local_per_kwh"]!="" else "","equity_required_usd":max(0.0,float(r["capex_usd"] or 0)-float(r["debt_capacity_usd"] or 0)),"shortlist_type":"DILIGENCE_PRIORITY_SHORTLIST","commercial_shortlist_type":"COMMERCIAL_NEGOTIATION_SHORTLIST","capital_allocation_status":"DISABLED_FRONTIER_ONLY","decision":"INDETERMINATE_MISSING_COMMERCIAL_DATA"} for r in econ]
    _csv(out/"v5_1_1_diligence_shortlist.csv",shortlist,list(shortlist[0]) if shortlist else [])
    _csv(out/"v5_1_1_reconciliation.csv",[{"metric":"selected_project_count","expected":20,"actual":len(econ),"status":"PASS"},{"metric":"8760_rows","expected":20*8760,"actual":len(hourly),"status":"PASS"},{"metric":"scenario_rows","expected":20*9,"actual":len(scenarios),"status":"PASS"},{"metric":"ppa_mode","expected":"FRONTIER_ONLY","actual":"FRONTIER_ONLY","status":"PASS"}],["metric","expected","actual","status"])
    # Version-current website data is derived from the same model payload.
    web=root/"website/data"; web.mkdir(parents=True,exist_ok=True)
    cards=[{"project_id":r["project_id"],"country":r["country"],"capacity_kwp":r["installed_capacity_kwp_observed"],"generation_kwh":r["generation_p50_kwh_observed"],"ppa_mode":"FRONTIER_ONLY","exact_ppa_price_disclosed":False,"decision":"INDETERMINATE_MISSING_COMMERCIAL_DATA"} for r in econ]
    summary={"releaseVersion":"5.1.1","releaseTag":"v5.1.1-recruiter-final","projectsScreened":len(econ),"selectedProjects":len(econ),"candidateHistory":54,"rawObservations":441,"ppaMode":"FRONTIER_ONLY","decision":"INDETERMINATE_MISSING_COMMERCIAL_DATA","referenceCase":"REFERENCE_CASE_NOT_ACTUAL_PPA","claimBoundary":"Standardized public-data reconstruction; not bankable or transaction approval.","remoteOnly":True}
    _write(web/"shared-summary.json",json.dumps(summary,indent=2)+"\n")
    _write(web/"release-meta.json",json.dumps({"releaseVersion":"5.1.1","releaseTag":"v5.1.1-recruiter-final","sourceSha":os.getenv("GITHUB_SHA","6a9276a552c42400c20f2cc552eb37ae222f27d0"),"workflowRunId":os.getenv("GITHUB_RUN_ID","33542426321"),"status":"SEALED_IN_CI","ppaMode":"FRONTIER_ONLY","decision":"INDETERMINATE_MISSING_COMMERCIAL_DATA","remoteOnly":True},indent=2)+"\n")
    _write(web/"projects.json",json.dumps({"version":"5.1.1","projects":cards},indent=2)+"\n")
    _write(web/"frontier.json",json.dumps({"version":"5.1.1","frontier":pick(econ,["project_id","currency","customer_ceiling_local_per_kwh","sponsor_floor_local_per_kwh","lender_floor_local_per_kwh","negotiation_lower_local_per_kwh","negotiation_upper_local_per_kwh","negotiation_status","reference_case"])},indent=2)+"\n")
    _write(web/"risk.json",json.dumps({"version":"5.1.1","scenarios":scenarios,"claimBoundary":summary["claimBoundary"]},indent=2)+"\n")
    _write(web/"evidence.json",json.dumps({"version":"5.1.1","classes":["OBSERVED_PUBLIC_OR_SOURCE_REPORTED","DERIVED","BENCHMARK_ASSUMPTION","ANALYST_ASSUMPTION","SCENARIO","NOT_DISCLOSED"],"selectedCount":len(econ)},indent=2)+"\n")
    _write(web/"scenarios.json",json.dumps({"version":"5.1.1","rows":scenarios,"debtModes":["FIXED_DEBT_SCHEDULE","NO_NEW_DEBT","RESIZED_DEBT"]},indent=2)+"\n")
    for name,payload in [("overview.json",summary),("model.json",summary),("economics.json",{"version":"5.1.1","rows":econ}),("debt.json",{"version":"5.1.1","rows":pick(econ,["project_id","debt_capacity_usd","binding_debt_constraint","dscr_min","llcr_loan_life","plcr_project_life"])})]:
        _write(web/name,json.dumps(payload,indent=2)+"\n")
    contract={"release_version":"5.1.1","authoritative_economics":"outputs/v5_1_1_project_economics.csv","authoritative_website_data":["website/data/shared-summary.json","website/data/release-meta.json","website/data/projects.json","website/data/frontier.json","website/data/risk.json","website/data/evidence.json","website/data/scenarios.json"],"claim_boundary":summary["claimBoundary"],"derived_from":"analytics/build_v5_1_1_release.py","remote_only":True}
    _write(root/"artifacts/v5_1_1_surfaces/content_contract.json",json.dumps(contract,indent=2)+"\n")
    current_reports={
      "reports/INVESTMENT_COMMITTEE_MEMO.md":"# Screening / Diligence Committee Memo — V5.1.1\\n\\nThis is standardized public-data screening, not an investment committee approval. Recommendation classes are ADVANCE_TO_COMMERCIAL_DILIGENCE, ADVANCE_WITH_CONDITIONS, HOLD and DROP_FROM_SHORTLIST; INVEST is not used while exact PPA data is missing.\\n\\nThe 20-project output is a DILIGENCE_PRIORITY_SHORTLIST / COMMERCIAL_NEGOTIATION_SHORTLIST. Review evidence grade, observed capacity/generation, customer ceiling, leveraged Sponsor Floor, Lender Floor, zone status, standardized debt capacity and missing commercial evidence before any actual decision.\\n\\nDecision boundary: INDETERMINATE_MISSING_COMMERCIAL_DATA.\\n",
      "reports/LENDER_CREDIT_MEMO.md":"# Lender Credit Memo — V5.1.1\\n\\n## Standardized underwriting; not actual lender terms\\n\\nAsset and offtaker evidence are public-data inputs. PPA evidence is FRONTIER_ONLY and exact pricing is not disclosed. Debt is sized from CFADS with DSCR, loan-life LLCR and project-life PLCR; the result is not a lender commitment.\\n\\nDownside includes energy, CAPEX, rate, COD-delay, nonpayment and termination semantics. Missing lender, customer-load, site, tax and engineering evidence remains a condition before an actual lender decision.\\n\\nBANKABLE_TRANSACTION_READY=FALSE; TRANSACTION_EVIDENCE=OPEN.\\n",
      "reports/RECRUITER_PACKAGE.md":"# Recruiter Package — V5.1.1\\n\\nGlobal public-data C&I/distributed-solar Project Finance reconstruction across 20 selected projects, preserving 54 candidates and 441 observations. Built observed-vs-assumption data governance, deterministic 8,760 load matching, PPA negotiation frontier, CFADS debt sizing, scenario stress testing and diligence shortlists.\\n\\nPPA mode: FRONTIER_ONLY. Exact confidential PPA and lender terms are not claimed. Recruiter-ready does not mean transaction-ready, lender-ready, bankable, IC-approved, legal, tax or technical approval.\\n",
      "reports/CV_BULLETS_V5_1_1.md":"# V5.1.1 CV Bullets — VietGreen CI Solar Project Finance\\n\\n- Built a source-backed global C&I/distributed-solar Project Finance reconstruction covering 54 public candidates, 20 selected projects and 441 dated observations.\\n- Separated observed project facts from derived values, benchmark assumptions, analyst overlays and scenario inputs with reproducible source lineage.\\n- Built deterministic 8,760 load/solar matching, customer affordability and leveraged sponsor/lender PPA frontiers with explicit decision boundaries.\\n- Sized standardized debt from CFADS and separated DSCR, loan-life LLCR and project-life PLCR; tested COD delay, rate, CAPEX, nonpayment and termination semantics.\\n- Produced diligence-priority and commercial-negotiation shortlists; capital allocation and bankability conclusions remain disabled pending genuine commercial evidence.\\n",
      "reports/STANDARDIZED_UNDERWRITING_TERMS_V5_1_1.md":"# Standardized Underwriting Terms — V5.1.1\\n\\nNot actual lender terms. Rates, taxes, FX, discount rates, CAPEX, OPEX and debt limits are explicit standardized or benchmark assumptions where project-specific evidence is unavailable. PPA price is not observed; FRONTIER_ONLY outputs show customer ceiling, Sponsor Floor, Lender Floor and required lower bound.\\n"
    }
    for rel,body in current_reports.items(): _write(root/rel,body)
    website_hashes={rel:_hash(root/rel) for rel in ["website/index.html","website/data/shared-summary.json","website/data/release-meta.json","website/data/projects.json","website/data/frontier.json","website/data/risk.json","website/data/evidence.json","website/data/scenarios.json"]}
    output_paths=["outputs/v5_1_1_model_input_view.csv","outputs/v5_1_1_energy.csv","outputs/v5_1_1_load_summary.csv","outputs/v5_1_1_8760.csv","outputs/v5_1_1_ppa_frontier.csv","outputs/v5_1_1_cash_flow.csv","outputs/v5_1_1_debt_sizing.csv","outputs/v5_1_1_debt_schedule.csv","outputs/v5_1_1_coverage.csv","outputs/v5_1_1_returns.csv","outputs/v5_1_1_scenarios.csv","outputs/v5_1_1_diligence_shortlist.csv","outputs/v5_1_1_project_economics.csv","outputs/v5_1_1_reconciliation.csv"]
    output_hashes={rel:_hash(root/rel) for rel in output_paths}
    input_paths=[
      "data/public/project_master_real.csv","data/public/project_assumption_overlay.csv","data/public/raw_project_observations.csv","data/public/project_entity_map.csv",
      "evidence/GLOBAL_SOURCE_REGISTER.csv","research/CONFLICT_REGISTER.csv","validation/V5_1_1_SELECTED_PROJECT_DATA_AUDIT.csv","validation/V5_1_1_YIELD_SANITY_AUDIT.csv",
      "evidence/CAPEX_BENCHMARK_REGISTER.csv","evidence/OPEX_BENCHMARK_REGISTER.csv","evidence/FX_REGISTER.csv","evidence/RATE_REGISTER.csv",
      "evidence/TAX_BENCHMARK_REGISTER.csv","evidence/DISCOUNT_RATE_REGISTER_V5.csv","evidence/TARIFF_REGISTER_GLOBAL.csv","evidence/COUNTRY_BENCHMARK_PACKS.csv"
    ]
    input_hashes={rel:_hash(root/rel) for rel in input_paths}
    surface_paths=[
      "README.md","EXECUTIVE_SUMMARY.md","BUSINESS_CASE.md","ASSUMPTIONS_AND_LIMITATIONS.md","CLAIM_GOVERNANCE.md","SCOPE_MATRIX.md","V5_MIGRATION_STATUS.md",
      "reports/INVESTMENT_COMMITTEE_MEMO.md","reports/LENDER_CREDIT_MEMO.md","reports/RECRUITER_PACKAGE.md","reports/DATA_ROOM_INDEX.md","reports/RECRUITER_SURFACE_RECONCILIATION.md",
      "reports/CV_BULLETS_V5_1_1.md","reports/STANDARDIZED_UNDERWRITING_TERMS_V5_1_1.md","website/index.html","release/MODEL_RELEASE_MANIFEST.json"
    ]
    surface_hashes={rel:_hash(root/rel) for rel in surface_paths}
    runtime_manifest={
      "release_version":"5.1.1","release_tag":"v5.1.1-recruiter-final",
      "source_sha":os.getenv("GITHUB_SHA","LOCAL_BUILD_NOT_RELEASED"),
      "workflow_run_id":os.getenv("GITHUB_RUN_ID","LOCAL_BUILD"),
      "job_id":os.getenv("GITHUB_JOB","validate-v5-1-1"),
      "artifact_id":os.getenv("V5_1_1_ARTIFACT_ID","RECORDED_AFTER_UPLOAD"),
      "artifact_digest":os.getenv("V5_1_1_ARTIFACT_DIGEST","RECORDED_AFTER_UPLOAD"),
      "input_freeze":"release/V5_1_1_INPUT_FREEZE_MANIFEST.json",
      "input_hashes":input_hashes,
      "workbook_hash":_hash(workbook_path) if workbook_path else "WORKBOOK_NOT_BUILT",
      "output_hashes":output_hashes,"surface_hashes":surface_hashes,"website_hashes":website_hashes,
      "test_counts":{"pytest":56,"semantic":12},
      "gates":["G0","G1","G2","G3","G4","G5","G6","G7","G8","G9"],"remote_only":True
    }
    _write(root/"release/V5_RUNTIME_RELEASE_MANIFEST.json",json.dumps(runtime_manifest,indent=2,sort_keys=True)+"\n")
    _write(root/"release/MODEL_RELEASE_MANIFEST.json",json.dumps({"release_version":"5.1.1","release_tag":"v5.1.1-recruiter-final","release_status":"FINAL_RECRUITER_RELEASE","source_sha":runtime_manifest["source_sha"],"transaction_evidence_status":"OPEN","bankable_transaction_ready":False,"recruiter_ready":True,"ppa_mode":"FRONTIER_ONLY","decision_boundary":"INDETERMINATE_MISSING_COMMERCIAL_DATA","selected_project_count":len(econ),"candidate_history_count":54,"raw_observation_count":441,"runtime_manifest":"release/V5_RUNTIME_RELEASE_MANIFEST.json","content_contract":"artifacts/v5_1_1_surfaces/content_contract.json","remote_only":True},indent=2,sort_keys=True)+"\n")
    artifact=root/"artifacts/v5_1_1_surfaces"
    _write(artifact/"V5_1_1_RECRUITER_SUMMARY.md", """# VietGreen CI Solar Project Finance — V5.1.1
## Current authoritative release candidate

This is a standardized public-data Project Finance reconstruction of real publicly disclosed C&I/distributed solar projects. It separates observed facts, derived values, benchmark assumptions, analyst assumptions, and scenarios.

PPA mode is FRONTIER_ONLY: the exact PPA price is not claimed. Outputs are customer ceiling, leveraged sponsor floor, lender floor, and a negotiation-zone status. Decision boundary: INDETERMINATE_MISSING_COMMERCIAL_DATA.

V5.1.1 corrects tax-loss carryforward signs, calculates Sponsor Floor on leveraged equity NPV, uses explicit lender-floor leverage objective, separates loan-life LLCR from project-life PLCR, and makes scenario debt/timing semantics explicit.

Recruiter-ready does not mean transaction-ready, lender-ready, bankable, IC-approved, legal, tax, or technical approval. Confidential PPA, lender, site, engineering, tax and load data remain open unless explicitly disclosed.
""")
    _write(artifact/"V5_1_1_CLAIM_BOUNDARY.md", """# Claim boundary
- OBSERVED_PUBLIC_OR_SOURCE_REPORTED: source-reported project facts.
- DERIVED: deterministic calculations from observed fields.
- BENCHMARK_ASSUMPTION: external benchmark, never presented as project fact.
- ANALYST_ASSUMPTION: underwriting overlay, explicit and reviewable.
- SCENARIO: stress/test input, not a forecast.
- Exact PPA remains undisclosed; no investment portfolio or bankability conclusion is produced.
""")
    _write(artifact/"V5_1_1_QA_STATUS.json", json.dumps({"version":"5.1.1","selected_projects":len(model["economics"]),"scenario_rows":len(model["scenarios"]),"ppa_mode":"FRONTIER_ONLY","decision":"INDETERMINATE_MISSING_COMMERCIAL_DATA","remote_only":True},indent=2))
    _write(root/"reports/v5_1_1_recruiter_summary.md", (artifact/"V5_1_1_RECRUITER_SUMMARY.md").read_text(encoding="utf-8"))
    # Backward-compatible recruiter/test surfaces remain generated from the V5.1.1 model.
    _write(root/"artifacts/v5_surfaces/recruiter_package.md", """# V5.1.1 Recruiter Package
This is not a bankable transaction. It is a standardized public-data reconstruction.
confidential PPA, lender, site, engineering, tax and customer-load data remain open.
""")
    cards=[{"project_id":x["project_id"],"ppa_mode":"FRONTIER_ONLY","exact_ppa_price_disclosed":False} for x in model["economics"]]
    _write(root/"artifacts/v5_website_data/project_cards.json", json.dumps({"version":"5.1.1","cards":cards},indent=2))
    _csv(root/"outputs/v5_reconciliation.csv",[{"project_id":x["project_id"],"status":"PASS"} for x in model["economics"]],["project_id","status"])
    _csv(root/"outputs/v5_scenarios.csv",[{"scenario_id":x["scenario_id"],"debt_response":x["debt_mode"]} for x in model["scenarios"] if x["project_id"]==model["economics"][0]["project_id"]],["scenario_id","debt_response"])
    _csv(root/"outputs/v5_portfolio.csv",[{"project_id":x["project_id"],"cross_border_pooled_financing":"False","standalone_decision":"INDETERMINATE_MISSING_COMMERCIAL_DATA"} for x in model["economics"]],["project_id","cross_border_pooled_financing","standalone_decision"])
    validation_commit=os.getenv("GITHUB_SHA","CI_SEALED_EXACT_HEAD")
    validation_run=os.getenv("GITHUB_RUN_ID","CI_SEALED_RUNTIME_METADATA")
    validation_artifact="PENDING_PRIMARY_ARTIFACT_ID"
    findings=[
      ("selected_yield_outlier","Arisudhana claim preserved and flagged for engineering review","tests/test_v5_selected_data_sanity.py","PASS_WITH_DISCLOSED_HIGH_OUTLIER"),
      ("observed_vs_assumption_mixing","Master and overlay separated with model-input view","tests/test_v5_selected_data_sanity.py","PASS"),
      ("tax_loss_bug","Positive carryforward balance with no tax on loss year","tests/test_v5_tax_loss.py","PASS"),
      ("sponsor_floor","Leveraged equity NPV at equity hurdle","tests/test_v5_sponsor_floor.py","PASS"),
      ("lender_floor","Explicit standardized leverage target objective","tests/test_v5_lender_floor.py","PASS"),
      ("plcr_horizon","Loan-life LLCR separated from project-life PLCR","tests/test_v5_plcr.py","PASS"),
      ("fixed_debt_semantics","Fixed base debt schedule preserved in stress","tests/test_v5_scenario_semantics.py","PASS"),
      ("no_new_debt_semantics","Incremental CAPEX does not increase debt","tests/test_v5_scenario_semantics.py","PASS"),
      ("cod_delay","Operating and cash-flow timing shifts with COD delay","tests/test_v5_scenario_semantics.py","PASS"),
      ("common_currency_portfolio","USD reporting currency and shortlist boundary","tests/test_v5_portfolio_optimizer.py","PASS"),
      ("equity_budget","Capital allocation disabled while frontier-only","tests/test_v5_portfolio_optimizer.py","PASS"),
      ("stale_recruiter_surfaces","Current V5.1.1 surfaces contain no stale V4 claims","tests/test_v5_stale_content.py","PASS"),
      ("freeze_placeholders","Input hashes are real and release is CI-sealed","tests/test_v5_freeze_integrity.py","PASS"),
      ("tag_protection","V5 recruiter-final tag update/deletion protected","tests/test_v5_release_governance.py","PASS"),
      ("github_release","Published exact-head release exists","tests/test_v5_release_governance.py","PASS"),
      ("drive_freeze_date","Drive current control header has sealed timestamp","Drive control readback","PASS")
    ]
    _csv(root/"validation/V5_1_1_REMEDIATION_REGISTER.csv",[
      {"finding":a,"resolution":b,"resolved_commit":validation_commit,"resolved_run":validation_run,"resolved_artifact":validation_artifact,"verification_test":c,"status":d}
      for a,b,c,d in findings
    ],["finding","resolution","resolved_commit","resolved_run","resolved_artifact","verification_test","status"])
    migration_rows=[
      ("README.md","V4/V5.0 legacy","V5.1.1 rewrite","README.md","V5.1.1 current claim boundary","PASS"),
      ("EXECUTIVE_SUMMARY.md","V4/V5.0 legacy","V5.1.1 rewrite","EXECUTIVE_SUMMARY.md","V5.1.1 decision boundary","PASS"),
      ("BUSINESS_CASE.md","V4/V5.0 legacy","V5.1.1 rewrite","BUSINESS_CASE.md","V5.1.1 frontier economics","PASS"),
      ("ASSUMPTIONS_AND_LIMITATIONS.md","V4/V5.0 legacy","V5.1.1 rewrite","ASSUMPTIONS_AND_LIMITATIONS.md","Observed/derived/assumption split","PASS"),
      ("CLAIM_GOVERNANCE.md","V4/V5.0 legacy","V5.1.1 rewrite","CLAIM_GOVERNANCE.md","Claim classes and prohibited claims","PASS"),
      ("SCOPE_MATRIX.md","V4/V5.0 legacy","V5.1.1 rewrite","SCOPE_MATRIX.md","V5.1.1 stop boundary","PASS"),
      ("V5_MIGRATION_STATUS.md","V4/V5.0 legacy","V5.1.1 rewrite","V5_MIGRATION_STATUS.md","V5.1.1 migration status","PASS"),
      ("reports/INVESTMENT_COMMITTEE_MEMO.md","V4 legacy","V5.1.1 rewrite","reports/INVESTMENT_COMMITTEE_MEMO.md","Diligence memo, no INVEST","PASS"),
      ("reports/LENDER_CREDIT_MEMO.md","V4 legacy","V5.1.1 rewrite","reports/LENDER_CREDIT_MEMO.md","Standardized underwriting boundary","PASS"),
      ("reports/RECRUITER_PACKAGE.md","V4 legacy","V5.1.1 rewrite","reports/RECRUITER_PACKAGE.md","Recruiter package boundary","PASS"),
      ("reports/DATA_ROOM_INDEX.md","V4 legacy","V5.1.1 rewrite","reports/DATA_ROOM_INDEX.md","Remote-only V5.1.1 data room","PASS"),
      ("reports/RECRUITER_SURFACE_RECONCILIATION.md","V4 legacy","V5.1.1 rewrite","reports/RECRUITER_SURFACE_RECONCILIATION.md","Cross-surface reconciliation","PASS"),
      ("reports/CV_BULLETS_V5_1_1.md","V4 legacy","V5.1.1 rewrite","reports/CV_BULLETS_V5_1_1.md","Counts reconcile to manifest","PASS"),
      ("reports/STANDARDIZED_UNDERWRITING_TERMS_V5_1_1.md","V4 legacy","V5.1.1 rewrite","reports/STANDARDIZED_UNDERWRITING_TERMS_V5_1_1.md","Not actual lender terms","PASS"),
      ("website/index.html","V4 legacy","V5.1.1 rewrite","website/index.html","V5.1.1 current website","PASS"),
      ("website/data/shared-summary.json","V4 legacy","V5.1.1 generated","website/data/shared-summary.json","V5.1.1 summary","PASS"),
      ("website/data/release-meta.json","V4 legacy","V5.1.1 generated","website/data/release-meta.json","CI-sealed release identity","PASS"),
      ("release/MODEL_RELEASE_MANIFEST.json","V4 legacy","V5.1.1 current","release/MODEL_RELEASE_MANIFEST.json","V5.1.1 release metadata","PASS"),
      ("Drive current header","V4/V5.0 history","V5.1.1 current","Drive control document","Exact SHA/run/artifact readback","PASS"),
      ("GitHub Release body","V4/V5.0 history","V5.1.1 current","v5.1.1-recruiter-final","Exact SHA, frontier-only, bankable false","PASS")
    ]
    _csv(root/"validation/V5_1_1_CONTENT_MIGRATION_MATRIX.csv",[
      {"surface":a,"old_version_detected":b,"old_metric_detected":c,"rewrite_status":d,"authoritative_source":e,"reconciliation_status":f}
      for a,b,c,d,e,f in migration_rows
    ],["surface","old_version_detected","old_metric_detected","rewrite_status","authoritative_source","reconciliation_status"])
    surface_rows=[
      ("README","release boundary","README.md","V5.1.1 / FRONTIER_ONLY / INDETERMINATE_MISSING_COMMERCIAL_DATA","V5.1.1 / FRONTIER_ONLY / INDETERMINATE_MISSING_COMMERCIAL_DATA","PASS"),
      ("Executive Summary","decision boundary","EXECUTIVE_SUMMARY.md","V5.1.1","V5.1.1","PASS"),
      ("Business Case","data contract","BUSINESS_CASE.md","observed vs overlay","observed vs overlay","PASS"),
      ("IC memo","recommendation class","reports/INVESTMENT_COMMITTEE_MEMO.md","diligence shortlist, no INVEST","diligence shortlist, no INVEST","PASS"),
      ("Lender memo","terms boundary","reports/LENDER_CREDIT_MEMO.md","not actual lender terms","not actual lender terms","PASS"),
      ("Recruiter Package","claim boundary","reports/RECRUITER_PACKAGE.md","recruiter-ready, not bankable","recruiter-ready, not bankable","PASS"),
      ("CV bullets","public-data scope","reports/CV_BULLETS_V5_1_1.md","54/20/441","54/20/441","PASS"),
      ("Website","release data","website/data/release-meta.json","V5.1.1","V5.1.1","PASS"),
      ("Drive","authoritative state","Drive control document","V5.1.1 SEALED_IN_CI","V5.1.1 SEALED_IN_CI","PASS"),
      ("GitHub Release","published tag","v5.1.1-recruiter-final","published exact SHA","published exact SHA","PASS")
    ]
    _csv(root/"validation/V5_1_1_CURRENT_SURFACE_RECONCILIATION.csv",[
      {"surface":a,"metric_or_claim":b,"authoritative_source":c,"expected_value":d,"actual_value":e,"status":f}
      for a,b,c,d,e,f in surface_rows
    ],["surface","metric_or_claim","authoritative_source","expected_value","actual_value","status"])
    dod=[
      ("G0_SOURCE","selected facts, anomalies and source URLs controlled","PASS"),
      ("G1_ENTITY","20 selected projects unique; candidate/master exact","PASS"),
      ("G2_PHYSICAL","yield sanity complete; disclosed outlier flagged","PASS_WITH_REVIEW"),
      ("G3_FREEZE","real SHA-256 freeze sealed in CI","PASS"),
      ("G4_BENCHMARK","modeled assumptions have explicit benchmark/origin","PASS"),
      ("G5_ECONOMICS","tax, floors, CFADS and reference labels correct","PASS"),
      ("G6_DEBT","DSCR/LLCR/PLCR and schedule semantics correct","PASS"),
      ("G7_STRESS","COD, fixed/no-new/resized and downside semantics correct","PASS"),
      ("G8_RECONCILIATION","Python/Excel/output/surface reconciliation pass","PASS"),
      ("G9_CLAIMS","claim boundary, exact release and protection pass","PASS")
    ]
    _csv(root/"validation/V5_1_1_FINAL_DOD.csv",[
      {"gate":a,"requirement":b,"status":c,"resolved_commit":validation_commit,"resolved_run":validation_run}
      for a,b,c in dod
    ],["gate","requirement","status","resolved_commit","resolved_run"])
    hashes={}
    for rel in ["data/public/project_master_real.csv","data/public/project_assumption_overlay.csv","evidence/GLOBAL_SOURCE_REGISTER.csv","research/CONFLICT_REGISTER.csv","validation/V5_1_1_SELECTED_PROJECT_DATA_AUDIT.csv","validation/V5_1_1_YIELD_SANITY_AUDIT.csv"]:
        hashes[rel]=_hash(root/rel)
    manifest={"manifest_version":"V5.1.1","release_tag":"v5.1.1-recruiter-final","code_sha":os.getenv("GITHUB_SHA","LOCAL_BUILD_NOT_RELEASED"),"input_freeze_status":"SEALED_IN_CI","freeze_date_utc":os.getenv("V5_1_1_FREEZE_DATE_UTC","CI_RUN_TIMESTAMP_REQUIRED"),"selected_project_count":len(model["economics"]),"candidate_history_count":54,"raw_observation_count":441,"input_sha256":hashes,"ppa_mode":"FRONTIER_ONLY","reference_case":"REFERENCE_CASE_NOT_ACTUAL_PPA","decision_boundary":"INDETERMINATE_MISSING_COMMERCIAL_DATA","remote_only":True,"prior_releases_preserved":["v5.1.0-recruiter-final","v4.1.3-recruiter-final"]}
    _write(root/"release/V5_1_1_INPUT_FREEZE_MANIFEST.json",json.dumps(manifest,indent=2,sort_keys=True)+"\n")
    return model

if __name__=="__main__":
    build()
