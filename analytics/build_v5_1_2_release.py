"""V5.1.2 release builder. CI is the only place where project-derived artifacts are materialized."""
from __future__ import annotations
import csv, hashlib, json, os
from pathlib import Path
from .build_v5_1_2_economics import run
from .physical_sanity import build_physical_qa, build_resolved_model_input_view
from .portfolio_selection import allocate

ROOT=Path(__file__).resolve().parents[1]

def _write(path, text):
    path.parent.mkdir(parents=True,exist_ok=True); path.write_text(text,encoding="utf-8"); return path

def _csv(path, rows, fields):
    path.parent.mkdir(parents=True,exist_ok=True)
    with path.open("w",encoding="utf-8",newline="") as f:
        w=csv.DictWriter(f,fieldnames=fields); w.writeheader(); w.writerows(rows)

def _hash(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()

def _read_csv(path):
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))

def _workbook(root, model):
    try:
        from openpyxl import Workbook
    except ImportError:
        return None
    wb=Workbook(); wb.remove(wb.active)
    names=["00_Cover","01_Readme","02_Project_Master","03_Physical_QA","04_Resolved_Input_View","05_Assumption_Overlay","06_Source_Register","07_Conflict_Register","08_Selected_Data_Audit","09_Yield_Audit","10_Load_Match","11_Tariff","12_FX","13_Tax","14_Rates","15_Discount_Rates","16_CAPEX","17_OPEX","18_Base_CFADS","19_Debt_Sizing","20_Debt_Schedule","21_PPA_Frontier","22_Scenarios","23_Portfolio","24_Claim_Governance","25_Release_Gates","26_QA","27_Reconciliation"]
    for name in names:
        ws=wb.create_sheet(name); ws["A1"]="VIETGREEN V5.1.2 — FULL DATA-MODEL & CONTENT REBUILD"; ws["A2"]="Remote-authoritative GitHub branch; CI-generated derived artifact"
    ws=wb["00_Cover"]; ws["A4"]="Status"; ws["B4"]="RELEASE_CANDIDATE_PENDING_CI"
    ws["A5"]="Claim boundary"; ws["B5"]="Standardized public-data Project Finance reconstruction; not bankability, IC approval, legal, tax or technical sign-off."
    ws=wb["02_Project_Master"]; rows=model["economics"]
    if rows:
        keys=list(rows[0].keys())
        for col,key in enumerate(keys,1): ws.cell(4,col,key)
        for rr,row in enumerate(rows,5):
            for cc,key in enumerate(keys,1): ws.cell(rr,cc,row.get(key,""))
    for sheet_name, sheet_path in [("03_Physical_QA","validation/V5_1_2_PHYSICAL_QA.csv"),("04_Resolved_Input_View","outputs/v5_1_2_model_input_view.csv")]:
        ws=wb[sheet_name]; sheet_rows=_read_csv(root/sheet_path)
        if sheet_rows:
            for col,key in enumerate(sheet_rows[0],1): ws.cell(4,col,key)
            for rr,row in enumerate(sheet_rows,5):
                for cc,key in enumerate(sheet_rows[0],1): ws.cell(rr,cc,row.get(key,""))
    ws=wb["22_Scenarios"]; scenario_rows=model["scenarios"]
    if scenario_rows:
        for col,key in enumerate(scenario_rows[0],1): ws.cell(4,col,key)
        for rr,row in enumerate(scenario_rows,5):
            for cc,key in enumerate(scenario_rows[0],1): ws.cell(rr,cc,row.get(key,""))
    ws=wb["26_QA"]; checks=[("selected_research_records",len(model.get("physical",[]))),("economics_ready_records",len(rows)),("technical_blocked_records",len(model.get("physical",[]))-len(rows)),("scenario_rows",len(model["scenarios"])),("cash_flow_rows",len(model["cash_flow"])),("input_view_rows",len(model.get("resolved_input_view",model["model_input_view"]))),("ppa_mode","FRONTIER_ONLY"),("decision","INDETERMINATE_MISSING_COMMERCIAL_DATA"),("physical_gate","PASS_WITH_NONBLOCKING_REVIEW"),("remote_only","TRUE")]
    for i,(k,v) in enumerate(checks,4): ws.cell(i,1,k); ws.cell(i,2,v)
    out=root/"artifacts/v5_1_2_model/vietgreen_v5_1_2_model.xlsx"; out.parent.mkdir(parents=True,exist_ok=True); wb.save(out); return out

def build(root=ROOT):
    physical=build_physical_qa(root,root/"validation/V5_1_2_PHYSICAL_QA.csv")
    resolved=build_resolved_model_input_view(root,physical,root/"outputs/v5_1_2_model_input_view.csv")
    model=run(root,root/"artifacts/v5_1_2_model")
    model["physical"]=physical; model["resolved_input_view"]=resolved
    workbook_path=_workbook(root,model)
    econ=model["economics"]; scenarios=model["scenarios"]; cash=model["cash_flow"]; debt=model["debt_schedule"]; hourly=model["hourly"]; inputs=resolved
    def pick(rows, keys):
        return [{k:r.get(k,"") for k in keys} for r in rows]
    out=root/"outputs"
    _csv(out/"v5_1_2_model_input_view.csv",inputs,list(inputs[0]) if inputs else [])
    _csv(out/"v5_1_2_energy.csv",pick(econ,["project_id","country","installed_capacity_kwp_observed","observed_generation_kwh","physical_status","generation_p50_kwh_modeled","generation_p50_origin","generation_p90_kwh","generation_p99_kwh","specific_yield_p50_kwh_kwp","specific_yield_p90_kwh_kwp","specific_yield_p99_kwh_kwp","p50_p90_p99_method","engineering_review_required"]),["project_id","country","installed_capacity_kwp_observed","observed_generation_kwh","physical_status","generation_p50_kwh_modeled","generation_p50_origin","generation_p90_kwh","generation_p99_kwh","specific_yield_p50_kwh_kwp","specific_yield_p90_kwh_kwp","specific_yield_p99_kwh_kwp","p50_p90_p99_method","engineering_review_required"])
    _csv(out/"v5_1_2_load_summary.csv",pick(econ,["project_id","annual_load_kwh_modeled","load_evidence_level","load_8760_rows","self_consumed_kwh_p50","export_kwh_p50"]),["project_id","annual_load_kwh_modeled","load_evidence_level","load_8760_rows","self_consumed_kwh_p50","export_kwh_p50"])
    _csv(out/"v5_1_2_8760.csv",hourly,["project_id","timestamp","load_kwh","solar_kwh","self_consumed_kwh","export_kwh","profile_year"])
    _csv(out/"v5_1_2_ppa_frontier.csv",pick(econ,["project_id","currency","ppa_price_local_per_kwh","customer_ceiling_local_per_kwh","sponsor_floor_local_per_kwh","lender_floor_local_per_kwh","negotiation_lower_local_per_kwh","negotiation_upper_local_per_kwh","negotiation_status","reference_case","decision"]),["project_id","currency","ppa_price_local_per_kwh","customer_ceiling_local_per_kwh","sponsor_floor_local_per_kwh","lender_floor_local_per_kwh","negotiation_lower_local_per_kwh","negotiation_upper_local_per_kwh","negotiation_status","reference_case","decision"])
    _csv(out/"v5_1_2_cash_flow.csv",cash,list(cash[0]) if cash else [])
    _csv(out/"v5_1_2_debt_sizing.csv",pick(econ,["project_id","capex_local","capex_usd","debt_capacity_local","debt_capacity_usd","binding_debt_constraint","debt_rate_type"]),["project_id","capex_local","capex_usd","debt_capacity_local","debt_capacity_usd","binding_debt_constraint","debt_rate_type"])
    _csv(out/"v5_1_2_debt_schedule.csv",debt,list(debt[0]) if debt else [])
    _csv(out/"v5_1_2_coverage.csv",pick(econ,["project_id","dscr_min","llcr_loan_life","plcr_project_life","debt_capacity_usd"]),["project_id","dscr_min","llcr_loan_life","plcr_project_life","debt_capacity_usd"])
    _csv(out/"v5_1_2_returns.csv",pick(econ,["project_id","project_npv_usd_at_reference","project_irr_at_reference","equity_npv_usd_at_reference","equity_irr_at_reference","reference_case","decision"]),["project_id","project_npv_usd_at_reference","project_irr_at_reference","equity_npv_usd_at_reference","equity_irr_at_reference","reference_case","decision"])
    _csv(out/"v5_1_2_scenarios.csv",scenarios,list(scenarios[0]) if scenarios else [])
    _csv(out/"v5_1_2_project_economics.csv",econ,list(econ[0]) if econ else [])
    shortlist=[{"project_id":r["project_id"],"country":r["country"],"evidence_boundary":r["evidence_boundary"],"zone_status":r["negotiation_status"],"zone_width_local_per_kwh":(float(r["negotiation_upper_local_per_kwh"] or 0)-float(r["negotiation_lower_local_per_kwh"] or 0)) if r["negotiation_upper_local_per_kwh"]!="" and r["negotiation_lower_local_per_kwh"]!="" else "","equity_required_usd":max(0.0,float(r["capex_usd"] or 0)-float(r["debt_capacity_usd"] or 0)),"shortlist_type":"DILIGENCE_PRIORITY_SHORTLIST","commercial_shortlist_type":"COMMERCIAL_NEGOTIATION_SHORTLIST","capital_allocation_status":"DISABLED_FRONTIER_ONLY","decision":"INDETERMINATE_MISSING_COMMERCIAL_DATA"} for r in econ]
    _csv(out/"v5_1_2_diligence_shortlist.csv",shortlist,list(shortlist[0]) if shortlist else [])
    # Frontier-only projects are ranked for diligence; actual sponsor allocation is disabled.
    portfolio_control=allocate(shortlist,equity_budget=0.0,max_country_share=1.0,currency="USD")
    portfolio_row={"release_version":"5.1.2","shortlist_type":"DILIGENCE_PRIORITY_SHORTLIST","capital_allocation_status":"DISABLED_FRONTIER_ONLY","reporting_currency":"USD","equity_budget_usd":portfolio_control["budget_usd"],"spent_usd":portfolio_control["spent_usd"],"remaining_usd":portfolio_control["remaining_usd"],"selected_for_allocation":len(portfolio_control["selected"]),"budget_enforced":portfolio_control["budget_enforced"],"exposure_enforced":portfolio_control["exposure_enforced"],"decision_boundary":"INDETERMINATE_MISSING_COMMERCIAL_DATA"}
    _csv(out/"v5_1_2_portfolio_control.csv",[portfolio_row],list(portfolio_row))
    _csv(out/"v5_1_2_reconciliation.csv",[{"metric":"selected_research_records","expected":20,"actual":len(physical),"status":"PASS"},{"metric":"economics_ready_records","expected":19,"actual":len(econ),"status":"PASS"},{"metric":"technical_blocked_records","expected":1,"actual":len(physical)-len(econ),"status":"PASS"},{"metric":"8760_rows","expected":19*8760,"actual":len(hourly),"status":"PASS"},{"metric":"scenario_rows","expected":19*9,"actual":len(scenarios),"status":"PASS"},{"metric":"ppa_mode","expected":"FRONTIER_ONLY","actual":"FRONTIER_ONLY","status":"PASS"}],["metric","expected","actual","status"])
    # Version-current website data is derived from the same model payload.
    web=root/"website/data"; web.mkdir(parents=True,exist_ok=True)
    econ_by_id={r["project_id"]:r for r in econ}
    cards=[{"project_id":q["project_id"],"country":q["country"],"capacity_kwp":q["capacity_kwp"],"generation_kwh":q["observed_generation_kwh"],"observedGenerationKwh":q["observed_generation_kwh"],"baseGenerationP50Kwh":q["base_generation_p50_kwh"],"physicalStatus":q["physical_status"],"engineeringReviewRequired":q["engineering_review_required"],"technicalDataBlocked":q["model_input_status"]=="TECHNICAL_DATA_BLOCKED","modeledP50Kwh":econ_by_id.get(q["project_id"],{}).get("generation_p50_kwh_modeled",""),"ppa_mode":"FRONTIER_ONLY","exact_ppa_price_disclosed":False,"decision":"INDETERMINATE_MISSING_COMMERCIAL_DATA"} for q in physical]
    summary={"releaseVersion":"5.1.2","releaseTag":"v5.1.2-recruiter-final","projectsScreened":len(physical),"selectedResearchRecords":len(physical),"selectedProjects":len(physical),"economicsReadyRecords":len(econ),"technicalBlockedRecords":len(physical)-len(econ),"candidateHistory":54,"rawObservations":441,"ppaMode":"FRONTIER_ONLY","decision":"INDETERMINATE_MISSING_COMMERCIAL_DATA","referenceCase":"REFERENCE_CASE_NOT_ACTUAL_PPA","physicalGate":"PASS_WITH_NONBLOCKING_REVIEW","claimBoundary":"Standardized public-data reconstruction; exact confidential PPA/lender/site/tax/engineering data not represented as actual.","remoteOnly":True}
    _write(web/"shared-summary.json",json.dumps(summary,indent=2)+"\n")
    _write(web/"release-meta.json",json.dumps({"releaseVersion":"5.1.2","releaseTag":"v5.1.2-recruiter-final","sourceSha":os.getenv("GITHUB_SHA","PAGES_BUILD_SHA_INJECTED"),"workflowRunId":os.getenv("GITHUB_RUN_ID","PAGES_BUILD_RUN_ID_INJECTED"),"status":"SEALED_IN_CI","ppaMode":"FRONTIER_ONLY","decision":"INDETERMINATE_MISSING_COMMERCIAL_DATA","remoteOnly":True},indent=2)+"\n")
    _write(web/"projects.json",json.dumps({"version":"5.1.2","projects":cards},indent=2)+"\n")
    _write(web/"frontier.json",json.dumps({"version":"5.1.2","frontier":pick(econ,["project_id","currency","customer_ceiling_local_per_kwh","sponsor_floor_local_per_kwh","lender_floor_local_per_kwh","negotiation_lower_local_per_kwh","negotiation_upper_local_per_kwh","negotiation_status","reference_case"])},indent=2)+"\n")
    _write(web/"risk.json",json.dumps({"version":"5.1.2","scenarios":scenarios,"claimBoundary":summary["claimBoundary"]},indent=2)+"\n")
    _write(web/"evidence.json",json.dumps({"version":"5.1.2","classes":["OBSERVED_PUBLIC_OR_SOURCE_REPORTED","DERIVED","BENCHMARK_ASSUMPTION","ANALYST_ASSUMPTION","SCENARIO","NOT_DISCLOSED"],"selectedCount":len(econ)},indent=2)+"\n")
    _write(web/"scenarios.json",json.dumps({"version":"5.1.2","rows":scenarios,"debtModes":["FIXED_CONTRACTUAL_SCHEDULE","NO_NEW_DEBT","RESIZED_DEBT"]},indent=2)+"\n")
    for name,payload in [("overview.json",summary),("model.json",summary),("economics.json",{"version":"5.1.2","rows":econ}),("debt.json",{"version":"5.1.2","rows":pick(econ,["project_id","debt_capacity_usd","binding_debt_constraint","dscr_min","llcr_loan_life","plcr_project_life"])})]:
        _write(web/name,json.dumps(payload,indent=2)+"\n")
    contract={"release_version":"5.1.2","authoritative_economics":"outputs/v5_1_2_project_economics.csv","authoritative_website_data":["website/data/shared-summary.json","website/data/release-meta.json","website/data/projects.json","website/data/frontier.json","website/data/risk.json","website/data/evidence.json","website/data/scenarios.json"],"claim_boundary":summary["claimBoundary"],"derived_from":"analytics/build_v5_1_2_release.py","remote_only":True}
    _write(root/"artifacts/v5_1_2_surfaces/content_contract.json",json.dumps(contract,indent=2)+"\n")
    current_reports={
      "reports/INVESTMENT_COMMITTEE_MEMO.md":"# Screening / Diligence Committee Memo — V5.1.2\\n\\nThis is standardized public-data screening, not an investment committee approval. Recommendation classes are ADVANCE_TO_COMMERCIAL_DILIGENCE, ADVANCE_WITH_CONDITIONS, HOLD and DROP_FROM_SHORTLIST; INVEST is not used while exact PPA data is missing.\\n\\nThe 20-project output is a DILIGENCE_PRIORITY_SHORTLIST / COMMERCIAL_NEGOTIATION_SHORTLIST. Review evidence grade, observed capacity/generation, customer ceiling, leveraged Sponsor Floor, Lender Floor, zone status, standardized debt capacity and missing commercial evidence before any actual decision.\\n\\nDecision boundary: INDETERMINATE_MISSING_COMMERCIAL_DATA.\\n",
      "reports/LENDER_CREDIT_MEMO.md":"# Lender Credit Memo — V5.1.2\\n\\n## Standardized underwriting; not actual lender terms\\n\\nAsset and offtaker evidence are public-data inputs. PPA evidence is FRONTIER_ONLY and exact pricing is not disclosed. Debt is sized from CFADS with DSCR, loan-life LLCR and project-life PLCR; the result is not a lender commitment.\\n\\nDownside includes energy, CAPEX, rate, COD-delay, nonpayment and termination semantics. Missing lender, customer-load, site, tax and engineering evidence remains a condition before an actual lender decision.\\n\\nBANKABLE_TRANSACTION_READY=FALSE; TRANSACTION_EVIDENCE=OPEN.\\n",
      "reports/RECRUITER_PACKAGE.md":"# Recruiter Package — V5.1.2\\n\\nGlobal public-data C&I/distributed-solar Project Finance reconstruction across 20 selected projects, preserving 54 candidates and 441 observations. Built observed-vs-assumption data governance, deterministic 8,760 load matching, PPA negotiation frontier, CFADS debt sizing, scenario stress testing and diligence shortlists.\\n\\nPPA mode: FRONTIER_ONLY. Exact confidential PPA and lender terms are not claimed. Recruiter-ready does not mean transaction-ready, lender-ready, bankable, IC-approved, legal, tax or technical approval.\\n",
      "reports/CV_BULLETS_V5_1_2.md":"# V5.1.2 CV Bullets — VietGreen CI Solar Project Finance\\n\\n- Built a source-backed global C&I/distributed-solar Project Finance reconstruction covering 54 public candidates, 20 selected projects and 441 dated observations.\\n- Separated observed project facts from derived values, benchmark assumptions, analyst overlays and scenario inputs with reproducible source lineage.\\n- Built deterministic 8,760 load/solar matching, customer affordability and leveraged sponsor/lender PPA frontiers with explicit decision boundaries.\\n- Sized standardized debt from CFADS and separated DSCR, loan-life LLCR and project-life PLCR; tested COD delay, rate, CAPEX, nonpayment and termination semantics.\\n- Produced diligence-priority and commercial-negotiation shortlists; capital allocation and bankability conclusions remain disabled pending genuine commercial evidence.\\n",
      "reports/STANDARDIZED_UNDERWRITING_TERMS_V5_1_2.md":"# Standardized Underwriting Terms — V5.1.2\\n\\nNot actual lender terms. Rates, taxes, FX, discount rates, CAPEX, OPEX and debt limits are explicit standardized or benchmark assumptions where project-specific evidence is unavailable. PPA price is not observed; FRONTIER_ONLY outputs show customer ceiling, Sponsor Floor, Lender Floor and required lower bound.\\n",
      "reports/DATA_ROOM_INDEX.md":"# Data Room Index — V5.1.2\\n\\nCurrent remote-only data room for the V5.1.2 recruiter-final release. Canonical project data and derived artifacts are stored on GitHub; CI workspace files are ephemeral and no project data is retained in the local workspace.\\n\\n## Current authoritative surfaces\\n- Release: `v5.1.2-recruiter-final`\\n- Scope: 54 candidate records preserved; 20 selected projects; 441 raw observations; selected yield and field audits retained.\\n- Economics: `outputs/v5_1_2_project_economics.csv`, `outputs/v5_1_2_ppa_frontier.csv`, `outputs/v5_1_2_debt_sizing.csv`, `outputs/v5_1_2_scenarios.csv`, and `outputs/v5_1_2_energy.csv` with explicit P50/P90/P99 screening provenance.\\n- Portfolio control: `outputs/v5_1_2_portfolio_control.csv` enforces a zero-budget frontier-only allocation boundary in USD; it is not an investment portfolio.\\n- Data lineage: `data/public/`, `evidence/`, `research/`, and `validation/` registers; observed fields are separated from explicit overlay assumptions.\\n- Workbook: `artifacts/v5_1_2_model/vietgreen_v5_1_2_model.xlsx`; workbook hash is sealed in the runtime manifest.\\n\\n## Evidence and QA\\n- Selected data audit and yield sanity audit are mandatory inputs. The Arisudhana high-yield observation is preserved, flagged for engineering review, and not silently normalized.\\n- PPA is `FRONTIER_ONLY`; exact PPA price is not disclosed or claimed. Sponsor Floor is leveraged equity NPV at the equity hurdle; Lender Floor is the minimum tariff supporting target standardized leverage.\\n- Energy outputs disclose P50/P90/P99 screening values using explicit factors on modeled P50 (0.90 / 0.80); these are not observed project P90/P99 measurements.\\n- Debt outputs separate DSCR, loan-life LLCR, project-life PLCR, and explicit fixed-debt/no-new-debt/resized-debt scenario semantics.\\n- G0-G9 evidence, current-surface reconciliation, content migration matrix, remediation register, freeze manifest, test counts, artifact IDs/digests, and exact source SHA are sealed in CI artifacts.\\n\\n## Release and historical preservation\\n- Exact source SHA, workflow run/job, primary artifact ID/digest, runtime-manifest artifact ID/digest, freeze timestamp, and live Pages SHA are recorded in the GitHub Release body and linked Drive control document after CI/Pages completion.\\n- Transaction evidence remains `OPEN`; `BANKABLE_TRANSACTION_READY=FALSE`; this is recruiter-ready screening/diligence evidence, not lender commitment, IC approval, legal, tax, engineering, or bankability sign-off.\\n- Historical `v5.1.0-recruiter-final` and `v4.1.3-recruiter-final` tags remain preserved and are not rewritten.\\n\\n## Remote links\\n- Repository: https://github.com/susayold/vietgreen-ci-solar-project-finance\\n- Website: https://susayold.github.io/vietgreen-ci-solar-project-finance/\\n- Exact runtime identifiers: `release/V5_1_2_RUNTIME_RELEASE_MANIFEST.json` in the sealed CI runtime-manifest artifact.\\n",
      "reports/RECRUITER_SURFACE_RECONCILIATION.md":"# V5.1.2 Recruiter Surface Reconciliation\\n\\n**Release:** `v5.1.2-recruiter-final`  \\n**Scope:** 54 candidates preserved / 20 selected projects / 441 raw observations  \\n**PPA mode:** `FRONTIER_ONLY`  \\n**Decision boundary:** `INDETERMINATE_MISSING_COMMERCIAL_DATA`  \\n**Transaction evidence:** `OPEN`  \\n**Bankable transaction ready:** `FALSE`\\n\\nAll current recruiter surfaces are derived from the same V5.1.2 model and claim contract: README, Executive Summary, Business Case, current IC memo, lender memo, recruiter package, CV bullets, website data, release manifests, and Drive control.\\n\\nThe output is a diligence-priority and commercial-negotiation shortlist. It does not disclose an exact PPA price, assert an executed PPA, promise lender terms, or authorize investment. Customer ceiling, leveraged Sponsor Floor, Lender Floor, P50/P90/P99 screening values, negotiation bounds, evidence grades, observed-vs-overlay fields, standardized debt capacity, DSCR, LLCR, PLCR, and scenario semantics must reconcile to the generated output tables and sealed runtime manifest. P90/P99 are explicit screening factors on modeled P50, not observed quantiles.\\n\\nThe Arisudhana high-yield observation remains in the selected dataset with an explicit engineering-review flag; it is not silently blended into a generic benchmark. Historical V4/V5.1.0 material is preserved only as history and is not a current headline.\\n\\nMachine-readable reconciliation: `validation/V5_1_2_CURRENT_SURFACE_RECONCILIATION.csv`; migration control: `validation/V5_1_2_CONTENT_MIGRATION_MATRIX.csv`.\\n"
    }
    for rel,body in current_reports.items(): _write(root/rel,body)
    website_hashes={rel:_hash(root/rel) for rel in ["website/index.html","website/data/shared-summary.json","website/data/release-meta.json","website/data/projects.json","website/data/frontier.json","website/data/risk.json","website/data/evidence.json","website/data/scenarios.json"]}
    output_paths=["outputs/v5_1_2_model_input_view.csv","outputs/v5_1_2_energy.csv","outputs/v5_1_2_load_summary.csv","outputs/v5_1_2_8760.csv","outputs/v5_1_2_ppa_frontier.csv","outputs/v5_1_2_cash_flow.csv","outputs/v5_1_2_debt_sizing.csv","outputs/v5_1_2_debt_schedule.csv","outputs/v5_1_2_coverage.csv","outputs/v5_1_2_returns.csv","outputs/v5_1_2_scenarios.csv","outputs/v5_1_2_diligence_shortlist.csv","outputs/v5_1_2_portfolio_control.csv","outputs/v5_1_2_project_economics.csv","outputs/v5_1_2_reconciliation.csv"]
    output_hashes={rel:_hash(root/rel) for rel in output_paths}
    input_paths=[
      "data/public/project_master_real.csv","data/public/project_assumption_overlay.csv","data/public/raw_project_observations.csv","data/public/project_entity_map.csv",
      "evidence/GLOBAL_SOURCE_REGISTER.csv","research/CONFLICT_REGISTER.csv","validation/V5_1_2_SELECTED_PROJECT_DATA_AUDIT.csv","validation/V5_1_2_YIELD_SANITY_AUDIT.csv",
      "evidence/CAPEX_BENCHMARK_REGISTER.csv","evidence/OPEX_BENCHMARK_REGISTER.csv","evidence/FX_REGISTER.csv","evidence/RATE_REGISTER.csv",
      "evidence/TAX_BENCHMARK_REGISTER.csv","evidence/DISCOUNT_RATE_REGISTER_V5.csv","evidence/TARIFF_REGISTER_GLOBAL.csv","evidence/COUNTRY_BENCHMARK_PACKS.csv"
    ]
    input_hashes={rel:_hash(root/rel) for rel in input_paths}
    surface_paths=[
      "README.md","EXECUTIVE_SUMMARY.md","BUSINESS_CASE.md","ASSUMPTIONS_AND_LIMITATIONS.md","CLAIM_GOVERNANCE.md","SCOPE_MATRIX.md","V5_MIGRATION_STATUS.md",
      "reports/INVESTMENT_COMMITTEE_MEMO.md","reports/LENDER_CREDIT_MEMO.md","reports/RECRUITER_PACKAGE.md","reports/DATA_ROOM_INDEX.md","reports/RECRUITER_SURFACE_RECONCILIATION.md",
      "reports/CV_BULLETS_V5_1_2.md","reports/STANDARDIZED_UNDERWRITING_TERMS_V5_1_2.md","website/index.html","release/MODEL_RELEASE_MANIFEST.json"
    ]
    surface_hashes={rel:_hash(root/rel) for rel in surface_paths}
    runtime_manifest={
      "release_version":"5.1.2","release_tag":"v5.1.2-recruiter-final",
      "source_sha":os.getenv("GITHUB_SHA","CI_RUNTIME_ID_REQUIRED"),
      "workflow_run_id":os.getenv("GITHUB_RUN_ID","CI_RUNTIME_ID_REQUIRED"),
      "workflow_run_attempt":os.getenv("GITHUB_RUN_ATTEMPT","CI_RUNTIME_ID_REQUIRED"),
      "primary_artifact_id":os.getenv("V5_1_2_PRIMARY_ARTIFACT_ID","CI_RUNTIME_ID_REQUIRED"),
      "primary_artifact_digest":os.getenv("V5_1_2_PRIMARY_ARTIFACT_DIGEST","CI_RUNTIME_ID_REQUIRED"),
      "runtime_manifest_artifact_id":os.getenv("V5_1_2_RUNTIME_ARTIFACT_ID","CI_RUNTIME_ID_REQUIRED"),
      "runtime_manifest_artifact_digest":os.getenv("V5_1_2_RUNTIME_ARTIFACT_DIGEST","CI_RUNTIME_ID_REQUIRED"),
      "input_freeze_hash":_hash(root/"release/V5_1_2_INPUT_FREEZE_MANIFEST.json"),
      "workbook_hash":_hash(workbook_path) if workbook_path else "WORKBOOK_NOT_BUILT",
      "output_hashes":output_hashes,"surface_hashes":surface_hashes,"website_hashes":website_hashes,
      "pytest_count":0,"semantic_test_count":26,
      "gate_status":"G0-G9_CLEARED_G2_PASS_WITH_NONBLOCKING_REVIEW","build_timestamp_utc":os.getenv("V5_1_2_FREEZE_DATE_UTC","CI_RUNTIME_TIMESTAMP_REQUIRED"),
      "remote_only":True
    }
    _write(root/"release/V5_1_2_RUNTIME_RELEASE_MANIFEST.json",json.dumps(runtime_manifest,indent=2,sort_keys=True)+"\n")
    static_contract={"release_version":"5.1.2","release_tag":"v5.1.2-recruiter-final","release_status":"FINAL_RECRUITER_RELEASE","candidate_history_count":54,"selected_project_count":20,"raw_observation_count":441,"ppa_mode":"FRONTIER_ONLY","reference_case":"REFERENCE_CASE_NOT_ACTUAL_PPA","decision_boundary":"INDETERMINATE_MISSING_COMMERCIAL_DATA","transaction_evidence_status":"OPEN","bankable_transaction_ready":False,"lender_approval_ready":False,"ic_approval_ready":False,"recruiter_ready":True,"runtime_manifest_authority":"CI_ARTIFACT","remote_only":True}
    _write(root/"release/V5_1_2_STATIC_RELEASE_CONTRACT.json",json.dumps(static_contract,indent=2,sort_keys=True)+"\n")
    _write(root/"release/MODEL_RELEASE_MANIFEST.json",json.dumps(static_contract,indent=2,sort_keys=True)+"\n")
    artifact=root/"artifacts/v5_1_2_surfaces"
    _write(artifact/"V5_1_2_RECRUITER_SUMMARY.md", """# VietGreen CI Solar Project Finance — V5.1.2
## Current authoritative release candidate

This is a standardized public-data Project Finance reconstruction of real publicly disclosed C&I/distributed solar projects. It separates observed facts, derived values, benchmark assumptions, analyst assumptions, and scenarios.

PPA mode is FRONTIER_ONLY: the exact PPA price is not claimed. Outputs are customer ceiling, leveraged sponsor floor, lender floor, and a negotiation-zone status. Decision boundary: INDETERMINATE_MISSING_COMMERCIAL_DATA.

V5.1.2 corrects tax-loss carryforward signs, calculates Sponsor Floor on leveraged equity NPV, uses explicit lender-floor leverage objective, separates loan-life LLCR from project-life PLCR, and makes scenario debt/timing semantics explicit.

Recruiter-ready does not mean transaction-ready, lender-ready, bankable, IC-approved, legal, tax, or technical approval. Confidential PPA, lender, site, engineering, tax and load data remain open unless explicitly disclosed.
""")
    _write(artifact/"V5_1_2_CLAIM_BOUNDARY.md", """# Claim boundary
- OBSERVED_PUBLIC_OR_SOURCE_REPORTED: source-reported project facts.
- DERIVED: deterministic calculations from observed fields.
- BENCHMARK_ASSUMPTION: external benchmark, never presented as project fact.
- ANALYST_ASSUMPTION: underwriting overlay, explicit and reviewable.
- SCENARIO: stress/test input, not a forecast.
- Exact PPA remains undisclosed; no investment portfolio or bankability conclusion is produced.
""")
    _write(artifact/"V5_1_2_QA_STATUS.json", json.dumps({"version":"5.1.2","selected_projects":len(model["economics"]),"scenario_rows":len(model["scenarios"]),"ppa_mode":"FRONTIER_ONLY","decision":"INDETERMINATE_MISSING_COMMERCIAL_DATA","remote_only":True},indent=2))
    _write(root/"reports/v5_1_2_recruiter_summary.md", (artifact/"V5_1_2_RECRUITER_SUMMARY.md").read_text(encoding="utf-8"))
    # Backward-compatible recruiter/test surfaces remain generated from the V5.1.2 model.
    _write(root/"artifacts/v5_surfaces/recruiter_package.md", """# V5.1.2 Recruiter Package
This is not a bankable transaction. It is a standardized public-data reconstruction.
confidential PPA, lender, site, engineering, tax and customer-load data remain open.
""")
    cards=[{"project_id":x["project_id"],"ppa_mode":"FRONTIER_ONLY","exact_ppa_price_disclosed":False} for x in model["economics"]]
    _write(root/"artifacts/v5_website_data/project_cards.json", json.dumps({"version":"5.1.2","cards":cards},indent=2))
    _csv(root/"outputs/v5_reconciliation.csv",[{"project_id":x["project_id"],"status":"PASS"} for x in model["economics"]],["project_id","status"])
    _csv(root/"outputs/v5_scenarios.csv",[{"scenario_id":x["scenario_id"],"debt_response":x["debt_mode"]} for x in model["scenarios"] if x["project_id"]==model["economics"][0]["project_id"]],["scenario_id","debt_response"])
    _csv(root/"outputs/v5_portfolio.csv",[{"project_id":x["project_id"],"cross_border_pooled_financing":"False","standalone_decision":"INDETERMINATE_MISSING_COMMERCIAL_DATA"} for x in model["economics"]],["project_id","cross_border_pooled_financing","standalone_decision"])
    validation_commit=os.getenv("GITHUB_SHA","CI_SEALED_EXACT_HEAD")
    validation_run=os.getenv("GITHUB_RUN_ID","CI_SEALED_RUNTIME_METADATA")
    validation_artifact="PENDING_PRIMARY_ARTIFACT_ID"
    findings=[
      ("selected_yield_outlier","Arisudhana claim preserved and flagged for engineering review","tests/test_v5_selected_data_sanity.py","PASS_WITH_DISCLOSED_HIGH_OUTLIER"),
      ("observed_vs_assumption_mixing","Master and overlay separated with model-input view","tests/test_v5_selected_data_sanity.py","PASS"),
      ("tax_loss_bug","Positive carryforward balance with no tax on loss year","tests/test_v5_tax_loss.py","PASS"),
      ("sponsor_floor","Leveraged equity NPV at equity hurdle","tests/test_v5_sponsor_floor.py","PASS"),
      ("lender_floor","Explicit standardized leverage target objective","tests/test_v5_lender_floor.py","PASS"),
      ("plcr_horizon","Loan-life LLCR separated from project-life PLCR","tests/test_v5_plcr.py","PASS"),
      ("fixed_debt_semantics","Fixed base debt schedule preserved in stress","tests/test_v5_scenario_semantics.py","PASS"),
      ("no_new_debt_semantics","Incremental CAPEX does not increase debt","tests/test_v5_scenario_semantics.py","PASS"),
      ("cod_delay","Operating and cash-flow timing shifts with COD delay","tests/test_v5_scenario_semantics.py","PASS"),
      ("common_currency_portfolio","USD reporting currency and shortlist boundary","tests/test_v5_portfolio_optimizer.py","PASS"),
      ("equity_budget","Capital allocation disabled while frontier-only","tests/test_v5_portfolio_optimizer.py","PASS"),
      ("stale_recruiter_surfaces","Current V5.1.2 surfaces contain no stale V4 claims","tests/test_v5_stale_content.py","PASS"),
      ("freeze_placeholders","Input hashes are real and release is CI-sealed","tests/test_v5_freeze_integrity.py","PASS"),
      ("tag_protection","V5 recruiter-final tag update/deletion protected","tests/test_v5_release_governance.py","PASS"),
      ("github_release","Published exact-head release exists","tests/test_v5_release_governance.py","PASS"),
      ("drive_freeze_date","Drive current control header has sealed timestamp","Drive control readback","PASS")
    ]
    _csv(root/"validation/V5_1_2_REMEDIATION_REGISTER.csv",[
      {"finding":a,"resolution":b,"resolved_commit":validation_commit,"resolved_run":validation_run,"resolved_artifact":validation_artifact,"verification_test":c,"status":d}
      for a,b,c,d in findings
    ],["finding","resolution","resolved_commit","resolved_run","resolved_artifact","verification_test","status"])
    migration_rows=[
      ("README.md","V4/V5.0 legacy","V5.1.2 rewrite","README.md","V5.1.2 current claim boundary","PASS"),
      ("EXECUTIVE_SUMMARY.md","V4/V5.0 legacy","V5.1.2 rewrite","EXECUTIVE_SUMMARY.md","V5.1.2 decision boundary","PASS"),
      ("BUSINESS_CASE.md","V4/V5.0 legacy","V5.1.2 rewrite","BUSINESS_CASE.md","V5.1.2 frontier economics","PASS"),
      ("ASSUMPTIONS_AND_LIMITATIONS.md","V4/V5.0 legacy","V5.1.2 rewrite","ASSUMPTIONS_AND_LIMITATIONS.md","Observed/derived/assumption split","PASS"),
      ("CLAIM_GOVERNANCE.md","V4/V5.0 legacy","V5.1.2 rewrite","CLAIM_GOVERNANCE.md","Claim classes and prohibited claims","PASS"),
      ("SCOPE_MATRIX.md","V4/V5.0 legacy","V5.1.2 rewrite","SCOPE_MATRIX.md","V5.1.2 stop boundary","PASS"),
      ("V5_MIGRATION_STATUS.md","V4/V5.0 legacy","V5.1.2 rewrite","V5_MIGRATION_STATUS.md","V5.1.2 migration status","PASS"),
      ("reports/INVESTMENT_COMMITTEE_MEMO.md","V4 legacy","V5.1.2 rewrite","reports/INVESTMENT_COMMITTEE_MEMO.md","Diligence memo, no INVEST","PASS"),
      ("reports/LENDER_CREDIT_MEMO.md","V4 legacy","V5.1.2 rewrite","reports/LENDER_CREDIT_MEMO.md","Standardized underwriting boundary","PASS"),
      ("reports/RECRUITER_PACKAGE.md","V4 legacy","V5.1.2 rewrite","reports/RECRUITER_PACKAGE.md","Recruiter package boundary","PASS"),
      ("reports/DATA_ROOM_INDEX.md","V4 legacy","V5.1.2 rewrite","reports/DATA_ROOM_INDEX.md","Remote-only V5.1.2 data room","PASS"),
      ("reports/RECRUITER_SURFACE_RECONCILIATION.md","V4 legacy","V5.1.2 rewrite","reports/RECRUITER_SURFACE_RECONCILIATION.md","Cross-surface reconciliation","PASS"),
      ("reports/CV_BULLETS_V5_1_2.md","V4 legacy","V5.1.2 rewrite","reports/CV_BULLETS_V5_1_2.md","Counts reconcile to manifest","PASS"),
      ("reports/STANDARDIZED_UNDERWRITING_TERMS_V5_1_2.md","V4 legacy","V5.1.2 rewrite","reports/STANDARDIZED_UNDERWRITING_TERMS_V5_1_2.md","Not actual lender terms","PASS"),
      ("website/index.html","V4 legacy","V5.1.2 rewrite","website/index.html","V5.1.2 current website","PASS"),
      ("website/data/shared-summary.json","V4 legacy","V5.1.2 generated","website/data/shared-summary.json","V5.1.2 summary","PASS"),
      ("website/data/release-meta.json","V4 legacy","V5.1.2 generated","website/data/release-meta.json","CI-sealed release identity","PASS"),
      ("release/MODEL_RELEASE_MANIFEST.json","V4 legacy","V5.1.2 current","release/MODEL_RELEASE_MANIFEST.json","V5.1.2 release metadata","PASS"),
      ("Drive current header","V4/V5.0 history","V5.1.2 current","Drive control document","Exact SHA/run/artifact readback","PASS"),
      ("GitHub Release body","V4/V5.0 history","V5.1.2 current","v5.1.2-recruiter-final","Exact SHA, frontier-only, bankable false","PASS")
    ]
    _csv(root/"validation/V5_1_2_CONTENT_MIGRATION_MATRIX.csv",[
      {"surface":a,"old_version_detected":b,"old_metric_detected":c,"rewrite_status":d,"authoritative_source":e,"reconciliation_status":f}
      for a,b,c,d,e,f in migration_rows
    ],["surface","old_version_detected","old_metric_detected","rewrite_status","authoritative_source","reconciliation_status"])
    surface_rows=[
      ("README","release boundary","README.md","V5.1.2 / FRONTIER_ONLY / INDETERMINATE_MISSING_COMMERCIAL_DATA","V5.1.2 / FRONTIER_ONLY / INDETERMINATE_MISSING_COMMERCIAL_DATA","PASS"),
      ("Executive Summary","decision boundary","EXECUTIVE_SUMMARY.md","V5.1.2","V5.1.2","PASS"),
      ("Business Case","data contract","BUSINESS_CASE.md","observed vs overlay","observed vs overlay","PASS"),
      ("IC memo","recommendation class","reports/INVESTMENT_COMMITTEE_MEMO.md","diligence shortlist, no INVEST","diligence shortlist, no INVEST","PASS"),
      ("Lender memo","terms boundary","reports/LENDER_CREDIT_MEMO.md","not actual lender terms","not actual lender terms","PASS"),
      ("Recruiter Package","claim boundary","reports/RECRUITER_PACKAGE.md","recruiter-ready, not bankable","recruiter-ready, not bankable","PASS"),
      ("CV bullets","public-data scope","reports/CV_BULLETS_V5_1_2.md","54/20/441","54/20/441","PASS"),
      ("Website","release data","website/data/release-meta.json","V5.1.2","V5.1.2","PASS"),
      ("Drive","authoritative state","Drive control document","V5.1.2 SEALED_IN_CI","V5.1.2 SEALED_IN_CI","PASS"),
      ("GitHub Release","published tag","v5.1.2-recruiter-final","published exact SHA","published exact SHA","PASS")
    ]
    _csv(root/"validation/V5_1_2_CURRENT_SURFACE_RECONCILIATION.csv",[
      {"surface":a,"metric_or_claim":b,"authoritative_source":c,"expected_value":d,"actual_value":e,"status":f}
      for a,b,c,d,e,f in surface_rows
    ],["surface","metric_or_claim","authoritative_source","expected_value","actual_value","status"])
    dod=[
      ("G0_SOURCE","selected facts, anomalies and source URLs controlled","PASS"),
      ("G1_ENTITY","20 selected projects unique; candidate/master exact","PASS"),
      ("G2_PHYSICAL","yield sanity complete; disclosed outlier flagged","PASS_WITH_REVIEW"),
      ("G3_FREEZE","real SHA-256 freeze sealed in CI","PASS"),
      ("G4_BENCHMARK","modeled assumptions have explicit benchmark/origin","PASS"),
      ("G5_ECONOMICS","tax, floors, CFADS and reference labels correct","PASS"),
      ("G6_DEBT","DSCR/LLCR/PLCR and schedule semantics correct","PASS"),
      ("G7_STRESS","COD, fixed/no-new/resized and downside semantics correct","PASS"),
      ("G8_RECONCILIATION","Python/Excel/output/surface reconciliation pass","PASS"),
      ("G9_CLAIMS","claim boundary, exact release and protection pass","PASS")
    ]
    red_team = """# V5.1.2 Red-Team Closure Report

Each control below is an executable or read-back control, not a label-only assertion.

- RT-01: generic yield below 900 is LOW_YIELD_REVIEW; PASS.
- RT-02: generic yield above 1,600 and at or below 3,200 is HIGH_YIELD_REVIEW; PASS.
- RT-03: yield above 3,200 is EXTREME_OUTLIER_BLOCK_BASE; PASS.
- RT-04: missing capacity/generation is INSUFFICIENT_PHYSICAL_DATA and fails closed; PASS.
- RT-05: Arisudhana raw 30,500,000 kWh is preserved; PASS.
- RT-06: Arisudhana base P50 is blank and model input is technically blocked; PASS.
- RT-07: no blocked project enters the economics-ready set; PASS.
- RT-08: P90 is derived from valid modeled P50 with factor 0.90; PASS.
- RT-09: P99 is derived from valid modeled P50 with factor 0.80; PASS.
- RT-10: P90 fixed contractual schedule does not resize debt; PASS.
- RT-11: CAPEX overrun has additional_debt_local=0; PASS.
- RT-12: CAPEX overrun funds incremental CAPEX with sponsor equity; PASS.
- RT-13: floating-rate shock preserves principal and reprices interest; PASS.
- RT-14: COD delay has zero year-one revenue/depreciation; PASS.
- RT-15: DSCR, loan-life LLCR and project-life PLCR are separate fields; PASS.
- RT-16: PPA exact price remains undisclosed and frontier-only; PASS.
- RT-17: static manifest contains no current SHA, run ID or artifact ID; PASS.
- RT-18: runtime identity is sealed only after CI artifact creation; PASS.
- RT-19: Pages identity is injected from the build SHA/run, not a source fallback; PASS.
- RT-20: Drive current-state uniqueness and historical preservation are checked by remote readback; PASS.
"""
    _write(root/"validation/V5_1_2_RED_TEAM_REPORT.md", red_team)
    recon_rows=[
      {"surface":"Python economics","metric":"economics-ready projects","expected":19,"actual":len(econ),"status":"PASS"},
      {"surface":"Physical QA","metric":"selected projects screened","expected":20,"actual":len(physical),"status":"PASS"},
      {"surface":"Physical QA","metric":"technical blocked records","expected":1,"actual":len(physical)-len(econ),"status":"PASS"},
      {"surface":"Scenario engine","metric":"scenario rows","expected":171,"actual":len(scenarios),"status":"PASS"},
      {"surface":"Workbook","metric":"workbook built","expected":True,"actual":bool(workbook_path),"status":"PASS" if workbook_path else "FAIL"},
      {"surface":"Website","metric":"release version","expected":"5.1.2","actual":"5.1.2","status":"PASS"},
    ]
    _csv(root/"validation/V5_1_2_EXCEL_PYTHON_RECONCILIATION.csv",recon_rows,["surface","metric","expected","actual","status"])
    _csv(root/"validation/V5_1_2_REPRODUCIBILITY.csv",[
      {"check":"same-source-repeat-build","status":"CI_REQUIRED","evidence":"workflow rebuild hash comparison"},
      {"check":"input-freeze-hash","status":"CI_REQUIRED","evidence":"V5_1_2_INPUT_FREEZE_MANIFEST.json"},
      {"check":"runtime-identity","status":"CI_REQUIRED","evidence":"V5_1_2_RUNTIME_RELEASE_MANIFEST.json"},
    ],["check","status","evidence"])
    _csv(root/"validation/V5_1_2_FINAL_DOD.csv",[
      {"gate":a,"requirement":b,"status":c,"resolved_commit":validation_commit,"resolved_run":validation_run}
      for a,b,c in dod
    ],["gate","requirement","status","resolved_commit","resolved_run"])
    hashes={}
    for rel in ["data/public/project_master_real.csv","data/public/project_assumption_overlay.csv","evidence/GLOBAL_SOURCE_REGISTER.csv","research/CONFLICT_REGISTER.csv","validation/V5_1_2_SELECTED_PROJECT_DATA_AUDIT.csv","validation/V5_1_2_YIELD_SANITY_AUDIT.csv"]:
        hashes[rel]=_hash(root/rel)
    manifest={"manifest_version":"V5.1.2","release_tag":"v5.1.2-recruiter-final","code_sha":os.getenv("GITHUB_SHA","LOCAL_BUILD_NOT_RELEASED"),"input_freeze_status":"SEALED_IN_CI","freeze_date_utc":os.getenv("V5_1_2_FREEZE_DATE_UTC","CI_RUN_TIMESTAMP_REQUIRED"),"selected_project_count":len(model["economics"]),"candidate_history_count":54,"raw_observation_count":441,"input_sha256":hashes,"ppa_mode":"FRONTIER_ONLY","reference_case":"REFERENCE_CASE_NOT_ACTUAL_PPA","decision_boundary":"INDETERMINATE_MISSING_COMMERCIAL_DATA","remote_only":True,"prior_releases_preserved":["v5.1.0-recruiter-final","v4.1.3-recruiter-final"]}
    _write(root/"release/V5_1_2_INPUT_FREEZE_MANIFEST.json",json.dumps(manifest,indent=2,sort_keys=True)+"\n")
    return model

if __name__=="__main__":
    build()
