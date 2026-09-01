import csv,json
from pathlib import Path
def test_v511_required_outputs_after_ci_build():
    names=["v5_1_1_model_input_view.csv","v5_1_1_energy.csv","v5_1_1_load_summary.csv","v5_1_1_8760.csv","v5_1_1_ppa_frontier.csv","v5_1_1_cash_flow.csv","v5_1_1_debt_sizing.csv","v5_1_1_debt_schedule.csv","v5_1_1_coverage.csv","v5_1_1_returns.csv","v5_1_1_scenarios.csv","v5_1_1_diligence_shortlist.csv","v5_1_1_project_economics.csv","v5_1_1_reconciliation.csv"]
    assert all((Path("outputs")/n).exists() for n in names)
    assert len(list(csv.DictReader((Path("outputs")/"v5_1_1_8760.csv").open(encoding="utf-8"))))==20*8760
def test_v511_workbook_and_manifests():
    from openpyxl import load_workbook
    wb=load_workbook("artifacts/v5_1_1_model/vietgreen_v5_1_1_model.xlsx",read_only=True)
    assert len(wb.sheetnames)==26 and wb.sheetnames[0]=="00_Cover" and wb.sheetnames[-1]=="25_QA"
    meta=json.loads(Path("release/V5_RUNTIME_RELEASE_MANIFEST.json").read_text())
    assert meta["release_version"]=="5.1.1" and meta["workbook_hash"]
